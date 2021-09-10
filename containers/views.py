from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.core.files.base import ContentFile
from openpyxl import load_workbook



from django.shortcuts import render
import openpyxl

def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

# you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        

        return render(request, 'index.html', {"excel_data":excel_data})



def upload_file(request):
    if request.method == 'POST':
        form = ModelFormWithFileField(request.POST, request.FILES)
        if form.is_valid():
            # file is saved
            workbook.save()
            return HttpResponseRedirect('/success/url/')
    else:
        form = ModelFormWithFileField()
    return render(request, 'upload.html', {'form': form})

from .models import Contract

def index (request):
    return render(request, 'containers/index.html')

class ContractList(ListView):
    model = Contract


class ContractDetail(DetailView):
    model = Contract


class ContractCreate(CreateView):
    model = Contract
    # Field must be same as the model attribute
    fields = ['name', 'date']
    success_url = reverse_lazy('contract_list')


class ContractUpdate(UpdateView):
    model = Contract
    # Field must be same as the model attribute
    fields = ['name', 'date']
    success_url = reverse_lazy('contract_list')


class ContractDelete(DeleteView):
    model = Contract
    success_url = reverse_lazy('contract_list')

